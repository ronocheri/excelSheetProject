import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def proccess_workbook(filename):

    wb=xl.load_workbook(filename)
    sheet=wb['Sheet1'] #get the first sheet

    title_cell = sheet.cell(1, 4)
    title_cell.value='updated_price'

    print(title_cell.value)
    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        print(cell.value)
        corrected_value=cell.value*0.9
        corrected_price_cell=sheet.cell(row,4)
        corrected_price_cell.value=corrected_value

    #create a refernce to data in column 4
    values=Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2') #add the chart and it's location
    wb.save(filename)

#calling to the function
proccess_workbook('transactions2.xlsx')